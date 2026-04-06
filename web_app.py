"""
web_app.py – Flask web interface for the Produce Order App.

Standalone version – no external database required. Store credentials and
inventory items are hardcoded. Orders are stored in a local JSON file for
persistence between restarts.

Run locally:
    python web_app.py

Or with Gunicorn:
    gunicorn web_app:app
"""

import io
import os
import json
import logging
import secrets
import smtplib
import ssl
from datetime import datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    jsonify,
    flash,
    send_file,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SECRET_KEY = os.environ.get("FLASK_SECRET_KEY", "change-me-in-production")

SMTP_HOST = os.environ.get("SMTP_HOST", "")
try:
    SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
except ValueError:
    SMTP_PORT = 587
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM_EMAIL = os.environ.get("SMTP_FROM_EMAIL", SMTP_USERNAME)

# Designated email address to receive order spreadsheets
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "")

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

# Path to persist orders as a JSON file (relative to working directory)
ORDERS_FILE = os.environ.get("ORDERS_FILE", "orders.json")

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

if not os.environ.get("ADMIN_PASSWORD"):
    logger.warning(
        "ADMIN_PASSWORD env var is not set – using insecure default. "
        "Set ADMIN_PASSWORD in your environment before deploying."
    )

# ---------------------------------------------------------------------------
# Hardcoded store credentials (no database required)
# ---------------------------------------------------------------------------
# Each entry: username -> {store_name, password, email}
STORES = {
    "scarborough_hannaford": {
        "store_name": "Scarborough Hannaford",
        "password": "Scarborough123!",
        "email": "",
    },
    "westbrook_hannaford": {
        "store_name": "Westbrook Hannaford",
        "password": "Westbrook123!",
        "email": "",
    },
    "riverside_hannaford": {
        "store_name": "Riverside Hannaford",
        "password": "Riverside123!",
        "email": "",
    },
    "rosemont_bakery": {
        "store_name": "Rosemont Bakery",
        "password": "Rosemont123!",
        "email": "",
    },
    "scratch_bakery": {
        "store_name": "Scratch Bakery",
        "password": "Scratch123!",
        "email": "",
    },
    "two_fat_cats": {
        "store_name": "Two Fat Cats Bakery",
        "password": "TwoFatCats123!",
        "email": "",
    },
    "beckys_diner": {
        "store_name": "Becky's Diner",
        "password": "Beckys123!",
        "email": "",
    },
}

# ---------------------------------------------------------------------------
# Hardcoded inventory placeholder
# ---------------------------------------------------------------------------
INVENTORY = [
    # Potatoes
    {"category": "Potatoes", "item": "White Chef - 50# bags", "qty": 100},
    {"category": "Potatoes", "item": "Yellow Chef - 50# bags", "qty": 100},
    {"category": "Potatoes", "item": "Red A - 50# bags", "qty": 100},
    {"category": "Potatoes", "item": "Red B - 50# bags", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 60 count", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 70 count", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 80 count", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 90 count", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 100 count", "qty": 100},
    {"category": "Potatoes", "item": "Russets - 50# boxes, 120 count", "qty": 100},
    # Apples
    {"category": "Apples", "item": "Macintosh - Loose Bulk 40#", "qty": 100},
    {"category": "Apples", "item": "Macintosh - 3# Bags in Case of 12", "qty": 100},
    {"category": "Apples", "item": "Cortland - Loose Bulk 40#", "qty": 100},
    {"category": "Apples", "item": "Cortland - 3# Bags in Case of 12", "qty": 100},
    {"category": "Apples", "item": "Honeycrisp - Loose Bulk 40#", "qty": 100},
    {"category": "Apples", "item": "Honeycrisp - 3# Bags in Case of 12", "qty": 100},
    # Onions
    {"category": "Onions", "item": "Red - 25# bags", "qty": 100},
    {"category": "Onions", "item": "Yellow - 25# bags", "qty": 100},
    # Eggs
    {"category": "Eggs", "item": "Loose Case - 15 dozen", "qty": 100},
    {"category": "Eggs", "item": "Retail Cartons Case - 15 dozen", "qty": 100},
    # Beets
    {"category": "Beets", "item": "Red - 20# bags", "qty": 100},
    {"category": "Beets", "item": "Candy Striped - 20# bags", "qty": 100},
    {"category": "Beets", "item": "Gold - 20# bags", "qty": 100},
]

# ---------------------------------------------------------------------------
# In-memory order store backed by a JSON file for persistence
# ---------------------------------------------------------------------------
_next_order_id: int = 1
_orders: list = []


def _load_orders() -> None:
    global _orders, _next_order_id
    if os.path.exists(ORDERS_FILE):
        try:
            with open(ORDERS_FILE, "r") as fh:
                data = json.load(fh)
            _orders = data.get("orders", [])
            _next_order_id = data.get("next_id", 1)
            logger.info("Loaded %d orders from %s", len(_orders), ORDERS_FILE)
        except Exception as exc:
            logger.error("Failed to load orders from file: %s", exc)
            _orders = []
            _next_order_id = 1


def _save_orders() -> None:
    try:
        with open(ORDERS_FILE, "w") as fh:
            json.dump({"orders": _orders, "next_id": _next_order_id}, fh, indent=2)
    except Exception as exc:
        logger.error("Failed to save orders to file: %s", exc)


def _add_order(
    store_name: str,
    category: str,
    item: str,
    qty: int,
    delivery_date: str,
    submitted_at: str,
    ordered_by: str,
) -> dict:
    global _next_order_id
    order = {
        "id": _next_order_id,
        "store_name": store_name,
        "category": category,
        "item": item,
        "qty": qty,
        "delivery_date": delivery_date,
        "submitted_at": submitted_at,
        "ordered_by": ordered_by,
    }
    _orders.append(order)
    _next_order_id += 1
    _save_orders()
    return order


def _get_orders(store_name: str = None, date_prefix: str = None) -> list:
    result = _orders
    if store_name:
        result = [o for o in result if o["store_name"] == store_name]
    if date_prefix:
        result = [o for o in result if o["delivery_date"].startswith(date_prefix)]
    return result


# Load persisted orders on startup
_load_orders()


# ---------------------------------------------------------------------------
# Allowed delivery dates (next 7 days starting tomorrow)
# ---------------------------------------------------------------------------

def _get_allowed_dates() -> list:
    """Return the next 7 days starting from tomorrow, e.g. 'April 7, 2026'."""
    today = datetime.utcnow().date()
    dates = []
    for i in range(1, 8):
        d = today + timedelta(days=i)
        # Build date string without zero-padded day (cross-platform)
        dates.append(f"{d.strftime('%B')} {d.day}, {d.year}")
    return dates


# ---------------------------------------------------------------------------
# Decorators
# ---------------------------------------------------------------------------

def login_required(f):
    """Redirect to login page when the store is not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "store_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Redirect to admin login page when admin is not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Email helpers
# ---------------------------------------------------------------------------

def _send_order_confirmation(
    to_email: str,
    store_name: str,
    ordered_by: str,
    delivery_date: str,
    items: list,
) -> None:
    """Send an order confirmation email to the store.

    Errors are logged but do not bubble up so that a missing/mis-configured
    SMTP server never blocks a successful order submission.
    """
    if not SMTP_HOST or not to_email:
        logger.info("Email not configured or no recipient – skipping confirmation email")
        return

    try:
        subject = f"Order Confirmation – {store_name}"

        lines = [
            f"Hi {store_name},",
            "",
            f"Your order has been placed successfully by {ordered_by}.",
            f"Delivery Date: {delivery_date}",
            "",
            "Order Summary:",
        ]
        for entry in items:
            lines.append(
                f"  • {entry.get('item', '')} ({entry.get('category', '')}) "
                f"– Qty: {entry.get('qty', 0)}"
            )
        lines += ["", "Thank you for your order!", "– Green Meadow Farms"]
        text_body = "\n".join(lines)

        item_rows = "".join(
            f"<tr><td>{entry.get('item', '')}</td><td>{entry.get('category', '')}</td>"
            f"<td style='text-align:center'>{entry.get('qty', 0)}</td></tr>"
            for entry in items
        )
        html_body = f"""
        <html><body>
        <p>Hi <strong>{store_name}</strong>,</p>
        <p>Your order has been placed successfully by <strong>{ordered_by}</strong>.</p>
        <p><strong>Delivery Date:</strong> {delivery_date}</p>
        <h3>Order Summary</h3>
        <table border="1" cellpadding="6" cellspacing="0"
               style="border-collapse:collapse;font-family:sans-serif">
          <thead>
            <tr style="background:#4a7c59;color:#fff">
              <th>Item</th><th>Category</th><th>Qty</th>
            </tr>
          </thead>
          <tbody>{item_rows}</tbody>
        </table>
        <br/>
        <p>Thank you for your order!<br/>– Green Meadow Farms</p>
        </body></html>
        """

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM_EMAIL
        msg["To"] = to_email
        msg.attach(MIMEText(text_body, "plain"))
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls(context=ssl.create_default_context())
            if SMTP_USERNAME and SMTP_PASSWORD:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_EMAIL, [to_email], msg.as_string())

        logger.info("Confirmation email sent to %s", to_email)
    except Exception as exc:
        logger.error("Failed to send confirmation email: %s", exc)


def _send_spreadsheet_email(spreadsheet: io.BytesIO, delivery_date: str) -> bool:
    """Email the order spreadsheet as an attachment to ADMIN_EMAIL.

    Returns True on success, False on failure.
    """
    if not SMTP_HOST:
        logger.info("SMTP not configured – cannot send spreadsheet email")
        return False
    if not ADMIN_EMAIL:
        logger.info("ADMIN_EMAIL not configured – cannot send spreadsheet email")
        return False

    try:
        subject = f"Order Spreadsheet – {delivery_date}"
        text_body = (
            f"Please find attached the order spreadsheet for delivery date: {delivery_date}.\n\n"
            "– Green Meadow Farms Order Portal"
        )

        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM_EMAIL
        msg["To"] = ADMIN_EMAIL
        msg.attach(MIMEText(text_body, "plain"))

        safe_date = delivery_date.replace(" ", "_").replace(",", "").replace("/", "-")
        filename = f"orders_{safe_date}.xlsx"
        part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part.set_payload(spreadsheet.getvalue())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls(context=ssl.create_default_context())
            if SMTP_USERNAME and SMTP_PASSWORD:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_EMAIL, [ADMIN_EMAIL], msg.as_string())

        logger.info("Spreadsheet emailed to %s", ADMIN_EMAIL)
        return True
    except Exception as exc:
        logger.error("Failed to send spreadsheet email: %s", exc)
        return False


# ---------------------------------------------------------------------------
# Spreadsheet builder
# ---------------------------------------------------------------------------

def _build_order_spreadsheet(orders: list, delivery_date: str) -> io.BytesIO:
    """Build an Excel workbook from a list of orders for the given delivery date.

    Layout:
      Row 1 (header): Category | Item | <Store 1> | <Store 2> | …
      Subsequent rows: one row per (category, item) pair, quantities per store.
    Items are sorted by category then name; stores are sorted alphabetically.
    """
    stores = sorted({o.get("store_name", "") for o in orders})
    items_set = set()
    for o in orders:
        items_set.add((o.get("category", ""), o.get("item", "")))
    items_sorted = sorted(items_set)

    qty_map: dict = {}
    for o in orders:
        key = (o.get("category", ""), o.get("item", ""), o.get("store_name", ""))
        qty_map[key] = qty_map.get(key, 0) + int(o.get("qty", 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = delivery_date[:10] if delivery_date else "Orders"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A7C59")
    category_fill = PatternFill("solid", fgColor="D9EAD3")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border_side = Side(style="thin", color="BBBBBB")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    headers = ["Category", "Item"] + stores
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    for row_idx, (category, item) in enumerate(items_sorted, start=2):
        cat_cell = ws.cell(row=row_idx, column=1, value=category)
        item_cell = ws.cell(row=row_idx, column=2, value=item)
        for cell in (cat_cell, item_cell):
            cell.fill = category_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")

        for col_idx, store in enumerate(stores, start=3):
            qty = qty_map.get((category, item, store), 0)
            qty_cell = ws.cell(row=row_idx, column=col_idx, value=qty if qty else "")
            qty_cell.alignment = center_align
            qty_cell.border = cell_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    for col_idx in range(3, 3 + len(stores)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        store_name = stores[col_idx - 3]
        ws.column_dimensions[col_letter].width = max(12, len(store_name) + 4)

    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    if "store_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if "store_id" in session:
        return redirect(url_for("dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            error = "Please enter both username and password."
        else:
            store = STORES.get(username)
            if store and secrets.compare_digest(password, store["password"]):
                session["store_id"] = username
                session["store_name"] = store["store_name"]
                session["store_email"] = store.get("email", "")
                logger.info("Store '%s' logged in", store["store_name"])
                return redirect(url_for("dashboard"))
            error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/logout-beacon", methods=["POST"])
def logout_beacon():
    """Endpoint called by navigator.sendBeacon() when the page/tab is closed."""
    if "store_id" in session:
        session.clear()
    return "", 204


@app.route("/dashboard")
@login_required
def dashboard():
    categories: dict = {}
    for item in INVENTORY:
        cat = item.get("category", "Other")
        categories.setdefault(cat, []).append(item)

    allowed_dates = _get_allowed_dates()
    cart = session.get("cart", [])
    cart_total = sum(int(i.get("qty", 0)) for i in cart)

    return render_template(
        "dashboard.html",
        store_name=session["store_name"],
        store_id=session["store_id"],
        categories=categories,
        allowed_dates=allowed_dates,
        cart=cart,
        cart_total=cart_total,
    )


# ---------------------------------------------------------------------------
# Cart API (session-based, JSON responses)
# ---------------------------------------------------------------------------

@app.route("/cart/add", methods=["POST"])
@login_required
def cart_add():
    data = request.get_json(silent=True) or {}
    category = data.get("category", "").strip()
    item = data.get("item", "").strip()

    try:
        qty = int(data.get("qty", 1))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid item data"}), 400

    if not category or not item or qty <= 0:
        return jsonify({"error": "Invalid item data"}), 400

    cart = session.get("cart", [])
    for entry in cart:
        if entry["category"] == category and entry["item"] == item:
            entry["qty"] = qty
            session["cart"] = cart
            return jsonify({"cart": cart, "total": sum(e["qty"] for e in cart)})

    cart.append({"category": category, "item": item, "qty": qty})
    session["cart"] = cart
    return jsonify({"cart": cart, "total": sum(e["qty"] for e in cart)})


@app.route("/cart/remove", methods=["POST"])
@login_required
def cart_remove():
    data = request.get_json(silent=True) or {}
    category = data.get("category", "").strip()
    item = data.get("item", "").strip()

    cart = session.get("cart", [])
    cart = [e for e in cart if not (e["category"] == category and e["item"] == item)]
    session["cart"] = cart
    return jsonify({"cart": cart, "total": sum(e["qty"] for e in cart)})


@app.route("/cart/clear", methods=["POST"])
@login_required
def cart_clear():
    session["cart"] = []
    return jsonify({"cart": [], "total": 0})


# ---------------------------------------------------------------------------
# Order submission
# ---------------------------------------------------------------------------

@app.route("/order/submit", methods=["POST"])
@login_required
def order_submit():
    data = request.get_json(silent=True) or {}
    delivery_date = data.get("delivery_date", "").strip()
    ordered_by = data.get("ordered_by", "").strip()
    cart = session.get("cart", [])

    if not delivery_date:
        return jsonify({"error": "Please select a delivery date."}), 400
    if not ordered_by:
        return jsonify({"error": "Please enter the name of the person placing the order."}), 400
    if not cart:
        return jsonify({"error": "Your cart is empty."}), 400

    submitted_at = datetime.utcnow().isoformat()
    store_name = session["store_name"]

    for item in cart:
        _add_order(
            store_name=store_name,
            category=item["category"],
            item=item["item"],
            qty=item["qty"],
            delivery_date=delivery_date,
            submitted_at=submitted_at,
            ordered_by=ordered_by,
        )

    _send_order_confirmation(
        to_email=session.get("store_email", ""),
        store_name=store_name,
        ordered_by=ordered_by,
        delivery_date=delivery_date,
        items=cart,
    )
    session["cart"] = []
    return jsonify({"success": True, "message": "Order submitted successfully!"})


# ---------------------------------------------------------------------------
# Order history
# ---------------------------------------------------------------------------

@app.route("/history")
@login_required
def history():
    store_name = session["store_name"]
    orders = _get_orders(store_name=store_name)
    orders = sorted(orders, key=lambda o: o.get("submitted_at", ""), reverse=True)

    return render_template(
        "history.html",
        store_name=store_name,
        orders=orders,
    )


# ---------------------------------------------------------------------------
# Admin – login / dashboard / spreadsheet download / email
# ---------------------------------------------------------------------------

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if session.get("is_admin"):
        return redirect(url_for("admin_dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user_ok = secrets.compare_digest(username, ADMIN_USERNAME)
        pass_ok = secrets.compare_digest(password, ADMIN_PASSWORD)
        if user_ok and pass_ok:
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        error = "Invalid admin credentials."

    return render_template("admin_login.html", error=error)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    allowed_dates = _get_allowed_dates()
    selected_date = request.args.get("delivery_date", "").strip()
    orders = []
    stores = []
    items_sorted = []
    qty_map: dict = {}

    if selected_date:
        orders = _get_orders(date_prefix=selected_date)
        stores = sorted({o.get("store_name", "") for o in orders})
        items_set = set()
        for o in orders:
            items_set.add((o.get("category", ""), o.get("item", "")))
        items_sorted = sorted(items_set)
        for o in orders:
            key = (o.get("category", ""), o.get("item", ""), o.get("store_name", ""))
            qty_map[key] = qty_map.get(key, 0) + int(o.get("qty", 0))

    return render_template(
        "admin.html",
        allowed_dates=allowed_dates,
        selected_date=selected_date,
        stores=stores,
        items=items_sorted,
        qty_map=qty_map,
        admin_email=ADMIN_EMAIL,
    )


@app.route("/admin/orders/spreadsheet")
@admin_required
def admin_spreadsheet():
    delivery_date = request.args.get("delivery_date", "").strip()
    if not delivery_date:
        flash("Please select a delivery date.", "error")
        return redirect(url_for("admin_dashboard"))

    orders = _get_orders(date_prefix=delivery_date)
    if not orders:
        flash("No orders found for that delivery date.", "warning")
        return redirect(url_for("admin_dashboard", delivery_date=delivery_date))

    spreadsheet = _build_order_spreadsheet(orders, delivery_date)
    safe_date = delivery_date.replace(" ", "_").replace(",", "").replace("/", "-")
    filename = f"orders_{safe_date}.xlsx"

    return send_file(
        spreadsheet,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/orders/email", methods=["POST"])
@admin_required
def admin_email_spreadsheet():
    """Email the order spreadsheet for the selected delivery date to ADMIN_EMAIL."""
    delivery_date = request.form.get("delivery_date", "").strip()
    if not delivery_date:
        flash("Please select a delivery date.", "error")
        return redirect(url_for("admin_dashboard"))

    orders = _get_orders(date_prefix=delivery_date)
    if not orders:
        flash("No orders found for that delivery date.", "warning")
        return redirect(url_for("admin_dashboard", delivery_date=delivery_date))

    spreadsheet = _build_order_spreadsheet(orders, delivery_date)
    ok = _send_spreadsheet_email(spreadsheet, delivery_date)
    if ok:
        flash(f"Spreadsheet emailed to {ADMIN_EMAIL}.", "success")
    else:
        flash(
            "Failed to send email. Check SMTP and ADMIN_EMAIL configuration.",
            "error",
        )
    return redirect(url_for("admin_dashboard", delivery_date=delivery_date))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
